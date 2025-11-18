"""
Data Export Module
Handles secure export to Excel and CSV formats
"""

import pandas as pd
from io import BytesIO
from datetime import datetime
from typing import Optional
import logging

logger = logging.getLogger(__name__)


class ExportError(Exception):
    """Custom export error"""
    pass


class DataExporter:
    """Export data to various formats"""
    
    @staticmethod
    def to_excel(
        df: pd.DataFrame,
        sheet_name: Optional[str] = None,
        include_timestamp: bool = True
    ) -> BytesIO:
        """
        Export DataFrame to Excel format with clickable hyperlinks
        
        Args:
            df: DataFrame to export
            sheet_name: Name for the Excel sheet
            include_timestamp: Whether to include timestamp in sheet name
            
        Returns:
            BytesIO object containing Excel file
            
        Raises:
            ExportError: If export fails
        """
        try:
            if df.empty:
                raise ExportError("Cannot export empty DataFrame")
            
            # Generate sheet name
            if sheet_name is None:
                sheet_name = datetime.now().strftime('%Y-%m-%d')
            elif include_timestamp:
                sheet_name = f"{sheet_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            
            # Ensure sheet name is valid (max 31 characters, no special chars)
            sheet_name = sheet_name[:31]
            sheet_name = ''.join(c for c in sheet_name if c.isalnum() or c in ['-', '_', ' '])
            
            output = BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Prepare DataFrame for export
                export_df = df.copy()
                
                # Keep URL columns for hyperlink creation but don't display them
                has_issue_url = 'Issue URL' in export_df.columns
                has_epic_url = 'Epic URL' in export_df.columns
                
                # Create display DataFrame (without URL columns)
                url_columns = []
                if has_issue_url:
                    url_columns.append('Issue URL')
                if has_epic_url:
                    url_columns.append('Epic URL')
                
                if url_columns:
                    display_columns = [col for col in export_df.columns if col not in url_columns]
                    display_df = export_df[display_columns]
                else:
                    display_df = export_df
                
                # Write to Excel
                display_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Get worksheet
                worksheet = writer.sheets[sheet_name]
                
                # Add hyperlinks to Epic/Story column if URLs are available
                if has_epic_url and 'Epic/Story' in export_df.columns:
                    # Find column index in display_df (not export_df)
                    epic_col_idx = list(display_df.columns).index('Epic/Story') + 1  # Excel is 1-indexed
                    
                    for row_idx, (epic_key, epic_url) in enumerate(zip(export_df['Epic/Story'], export_df['Epic URL']), start=2):
                        if epic_url and epic_key:  # Only add hyperlink if both exist
                            cell = worksheet.cell(row=row_idx, column=epic_col_idx)
                            cell.value = epic_key
                            cell.hyperlink = epic_url
                            cell.style = 'Hyperlink'
                
                # Add hyperlinks to Issue key column if URLs are available
                if has_issue_url and 'Issue key' in export_df.columns:
                    # Find column index in display_df (not export_df)
                    issue_key_col_idx = list(display_df.columns).index('Issue key') + 1  # Excel is 1-indexed
                    
                    for row_idx, (issue_key, issue_url) in enumerate(zip(export_df['Issue key'], export_df['Issue URL']), start=2):
                        if issue_url:
                            cell = worksheet.cell(row=row_idx, column=issue_key_col_idx)
                            cell.value = issue_key
                            cell.hyperlink = issue_url
                            cell.style = 'Hyperlink'
                
                # Auto-adjust column widths
                for idx, col in enumerate(display_df.columns, 1):
                    max_length = max(
                        display_df[col].astype(str).apply(len).max(),
                        len(col)
                    )
                    # Set column width (min 10, max 50)
                    worksheet.column_dimensions[
                        worksheet.cell(1, idx).column_letter
                    ].width = min(max(max_length + 2, 10), 50)
                
                # Format header row
                for cell in worksheet[1]:
                    cell.font = cell.font.copy(bold=True)
                    cell.fill = cell.fill.copy(fgColor="0052CC", patternType="solid")
                    from openpyxl.styles import Font, PatternFill
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="0052CC", end_color="0052CC", fill_type="solid")
            
            output.seek(0)
            logger.info(f"Successfully exported {len(df)} rows to Excel with hyperlinks")
            return output
            
        except Exception as e:
            logger.error(f"Excel export failed: {str(e)}")
            raise ExportError(f"Failed to export to Excel: {str(e)}")
    
    @staticmethod
    def to_csv(df: pd.DataFrame, include_index: bool = False) -> str:
        """
        Export DataFrame to CSV format
        
        Args:
            df: DataFrame to export
            include_index: Whether to include index in CSV
            
        Returns:
            CSV string
            
        Raises:
            ExportError: If export fails
        """
        try:
            if df.empty:
                raise ExportError("Cannot export empty DataFrame")
            
            csv_string = df.to_csv(index=include_index, encoding='utf-8')
            logger.info(f"Successfully exported {len(df)} rows to CSV")
            return csv_string
            
        except Exception as e:
            logger.error(f"CSV export failed: {str(e)}")
            raise ExportError(f"Failed to export to CSV: {str(e)}")
    
    @staticmethod
    def get_filename(
        base_name: str,
        extension: str,
        include_timestamp: bool = True
    ) -> str:
        """
        Generate safe filename
        
        Args:
            base_name: Base filename
            extension: File extension (with or without dot)
            include_timestamp: Whether to include timestamp
            
        Returns:
            Generated filename
        """
        # Ensure extension has dot
        if not extension.startswith('.'):
            extension = '.' + extension
        
        # Sanitize base name
        safe_name = ''.join(c for c in base_name if c.isalnum() or c in ['-', '_'])
        
        if include_timestamp:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{safe_name}_{timestamp}{extension}"
        else:
            filename = f"{safe_name}{extension}"
        
        return filename